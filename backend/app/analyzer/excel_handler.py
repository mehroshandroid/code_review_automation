import datetime
import re
from pathlib import Path

from openpyxl import load_workbook

PLACEHOLDER_PROJECT_NAME = "<project name>"
GENERAL_REMARKS_PREFIX = "general remarks"
REVIEWERS_LABEL = "reviewers"
DATED_LABEL = "dated"

# The real template (samplefiles/SampleCodeReview.xlsx) has no separate "Score"
# column: category rows carry an existing =AVERAGE(...) formula in the
# "Avg Points" column, and that same column is reused to hold each
# sub-criterion's raw score directly. Sub-criteria are not reliably labeled in
# the id column (the first sub-row under each category is often blank, and
# labels can be duplicated/typo'd) -- the category's own AVERAGE(...) formula
# range is the actual source of truth for how many rows belong to it, so
# sub-rows are matched positionally (N rows immediately following the
# category row, N = number of sub-criteria for that category) rather than by
# reading their id cell.
HEADER_ALIASES = {
    "id": ["clause", "category", "sub-criterion", "sub criterion", "criterion"],
    "avg_points": ["avg points", "average points"],
    "remarks": ["remarks"],
}

MAX_HEADER_SCAN_ROWS = 10


def aggregate_category_scores(sub_scores: dict) -> dict:
    values = [v["score"] for v in sub_scores.values() if v.get("score") is not None]
    if not values:
        avg_points = final_points = percent_points = None
    else:
        avg_points = round(sum(values) / len(values), 2)
        final_points = avg_points
        percent_points = round(final_points * 100, 1)
    return {
        "avg_points": avg_points,
        "final_points": final_points,
        "percent_points": percent_points,
        "sub_scores": sub_scores,
    }


def compute_total_score_pct(scores_by_category: dict) -> float | None:
    values = [
        result["percent_points"]
        for result in scores_by_category.values()
        if result.get("percent_points") is not None
    ]
    if not values:
        return None
    return round(sum(values) / len(values), 1)


def _normalize_id(value):
    """Normalize an id cell's value to the string form used as category_results keys.

    Excel-native numeric cells come back from openpyxl as int/float rather than
    str. Whole-number floats (e.g. 1.0) are formatted without the trailing
    ".0" so they match string keys like "1". A blank cell normalizes to None.
    """
    if value is None:
        return None
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _is_formula_cell(cell) -> bool:
    if cell.data_type == "f":
        return True
    return isinstance(cell.value, str) and cell.value.startswith("=")


def _set_cell(ws, row_idx: int, column: int, value) -> None:
    cell = ws.cell(row=row_idx, column=column)
    if _is_formula_cell(cell):
        return
    cell.value = value


def _find_header_row(ws) -> int:
    id_aliases = set(HEADER_ALIASES["id"])
    for row in ws.iter_rows(min_row=1, max_row=min(MAX_HEADER_SCAN_ROWS, ws.max_row)):
        for cell in row:
            if cell.value is not None and str(cell.value).strip().lower() in id_aliases:
                return cell.row
    raise ValueError(
        f"Could not find a header row (looked for one of {sorted(id_aliases)} "
        f"in the first {MAX_HEADER_SCAN_ROWS} rows)"
    )


def _resolve_columns(ws, header_row: int) -> dict:
    columns = {}
    for cell in ws[header_row]:
        if cell.value is None:
            continue
        header_text = str(cell.value).strip().lower()
        for key, aliases in HEADER_ALIASES.items():
            if header_text in aliases:
                columns[key] = cell.column
    missing = [key for key in HEADER_ALIASES if key not in columns]
    if missing:
        raise ValueError(f"Excel template missing expected columns: {missing}")
    return columns


def _iter_positional_sub_rows(ws, header_row: int, id_col: int, category_sub_ids: dict):
    """Yields (category_id, category_row, sub_id, sub_row) by walking the
    sheet positionally: for a row whose id cell matches a key in
    category_sub_ids, the next N rows (N = len(category_sub_ids[key])) are
    yielded in order as that category's sub-criteria, regardless of what
    their own id cells contain. Shared by populate_scores (writing) and
    extract_sub_criteria_descriptions (reading) so both stay in lockstep.
    """
    max_row = ws.max_row
    row = header_row + 1
    while row <= max_row:
        id_cell = ws.cell(row=row, column=id_col)
        category_id = _normalize_id(id_cell.value)
        if category_id in category_sub_ids:
            sub_ids = category_sub_ids[category_id]
            offset = 1
            for sub_id in sub_ids:
                sub_row = row + offset
                if sub_row > max_row:
                    return
                yield category_id, row, sub_id, sub_row
                offset += 1
            row += offset
        else:
            row += 1


AVERAGE_FORMULA_RE = re.compile(r"=AVERAGE\([A-Z]+(\d+):[A-Z]+(\d+)\)", re.IGNORECASE)


def discover_structure(ws) -> tuple[dict, dict]:
    """Discovers the full category/sub-criteria structure directly from the
    template, with no hardcoded platform-specific knowledge: a category row
    is any row whose Avg Points cell holds a pre-existing =AVERAGE(range)
    rollup formula (every real template's convention, regardless of
    platform); the formula's own row range tells us exactly how many
    sub-criterion rows follow and where. Sub-criterion ids are synthesized
    positionally ("{category_id}.{n}") since sub-row id cells are already
    known to be unreliable (blank/typo'd/duplicated) in real templates.
    Returns (categories, descriptions):
      categories: {category_id: {"name": str, "sub_criteria": [sub_id, ...]}}
      descriptions: {sub_id: str}
    """
    header_row = _find_header_row(ws)
    columns = _resolve_columns(ws, header_row)
    id_col = columns["id"]
    avg_col = columns["avg_points"]
    description_col = id_col + 1

    categories = {}
    descriptions = {}
    max_row = ws.max_row
    row = header_row + 1
    while row <= max_row:
        avg_cell = ws.cell(row=row, column=avg_col)
        match = AVERAGE_FORMULA_RE.match(str(avg_cell.value)) if _is_formula_cell(avg_cell) else None
        if match:
            category_id = _normalize_id(ws.cell(row=row, column=id_col).value)
            name_cell = ws.cell(row=row, column=description_col)
            category_name = str(name_cell.value).strip() if name_cell.value else ""
            start_row, end_row = int(match.group(1)), int(match.group(2))

            sub_ids = []
            for offset, sub_row in enumerate(range(start_row, end_row + 1), start=1):
                sub_id = f"{category_id}.{offset}"
                sub_ids.append(sub_id)
                desc_cell = ws.cell(row=sub_row, column=description_col)
                descriptions[sub_id] = str(desc_cell.value).strip() if desc_cell.value else ""

            categories[category_id] = {"name": category_name, "sub_criteria": sub_ids}
            row = end_row + 1
        else:
            row += 1
    return categories, descriptions


def populate_scores(ws, category_results: dict) -> None:
    """Writes sub-criterion scores/remarks into an already-loaded worksheet, in place."""
    header_row = _find_header_row(ws)
    columns = _resolve_columns(ws, header_row)
    id_col = columns["id"]
    score_col = columns["avg_points"]
    remarks_col = columns["remarks"]

    category_sub_ids = {cid: list(result["sub_scores"].keys()) for cid, result in category_results.items()}
    for category_id, _category_row, sub_id, sub_row in _iter_positional_sub_rows(ws, header_row, id_col, category_sub_ids):
        sub = category_results[category_id]["sub_scores"][sub_id]
        _set_cell(ws, sub_row, score_col, sub.get("score"))
        # Perfect scores are self-explanatory; only justify < 1 (or
        # not-evaluated) so the sheet reads as "here's what's wrong", not a
        # remark on every single row regardless of outcome. Explicitly clear
        # (not just skip) so a stale remark from a prior run against the same
        # template doesn't linger.
        remark = sub.get("remark") if sub.get("score") != 1 else None
        _set_cell(ws, sub_row, remarks_col, remark)


def extract_sub_criteria_descriptions(ws, categories: dict) -> dict:
    """Reads each sub-criterion's description text (the column immediately to
    the right of the id column) using the same positional row-matching as
    populate_scores, so the LLM prompt can be grounded in the template's
    actual wording instead of a bare sub-criterion id like "2.4" -- which the
    model otherwise has to guess the meaning of. Returns {sub_id: text},
    flattened across all categories. categories is CATEGORIES-shaped:
    {category_id: {"name": str, "sub_criteria": [sub_id, ...]}}.
    """
    header_row = _find_header_row(ws)
    columns = _resolve_columns(ws, header_row)
    id_col = columns["id"]
    description_col = id_col + 1

    category_sub_ids = {cid: cat["sub_criteria"] for cid, cat in categories.items()}
    descriptions = {}
    for _category_id, _category_row, sub_id, sub_row in _iter_positional_sub_rows(ws, header_row, id_col, category_sub_ids):
        desc_cell = ws.cell(row=sub_row, column=description_col)
        descriptions[sub_id] = str(desc_cell.value).strip() if desc_cell.value else ""
    return descriptions


def populate_metadata(
    ws,
    project_name: str,
    general_remarks: str,
    reviewer_name: str,
    review_date: datetime.date,
) -> None:
    """Fills in the template's whole-review metadata cells (title, general
    remarks, reviewer, date), located by label text search rather than fixed
    coordinates so it tolerates minor row/column drift between template
    versions. Labels are matched case-insensitively; unmatched labels are
    left untouched (non-fatal -- a template missing one of these is still
    usable for scoring).
    """
    for row in ws.iter_rows():
        for cell in row:
            if not isinstance(cell.value, str):
                continue
            text = cell.value.strip().lower()
            if text == PLACEHOLDER_PROJECT_NAME:
                _set_cell(ws, cell.row, cell.column, project_name)
            elif text.startswith(GENERAL_REMARKS_PREFIX):
                _set_cell(ws, cell.row, cell.column, f"General Remarks: {general_remarks}")
            elif text.startswith(REVIEWERS_LABEL):
                _set_cell(ws, cell.row, cell.column + 1, reviewer_name)
            elif text == DATED_LABEL:
                _set_cell(ws, cell.row, cell.column + 1, review_date)


def generate_review_excel(
    template_path: Path,
    output_path: Path,
    category_results: dict,
    project_name: str,
    general_remarks: str,
    reviewer_name: str = "Claude",
    review_date: datetime.date | None = None,
) -> None:
    wb = load_workbook(template_path)
    ws = wb.active
    populate_scores(ws, category_results)
    populate_metadata(
        ws,
        project_name=project_name,
        general_remarks=general_remarks,
        reviewer_name=reviewer_name,
        review_date=review_date or datetime.date.today(),
    )
    wb.save(output_path)
