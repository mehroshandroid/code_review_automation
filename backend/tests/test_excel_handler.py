import datetime
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font

from app.analyzer.excel_handler import (
    aggregate_category_scores,
    compute_total_score_pct,
    extract_sub_criteria_descriptions,
    generate_review_excel,
    populate_metadata,
    populate_scores,
)

FIXTURES_DIR = Path(__file__).parent / "fixtures"


def _build_template(path: Path) -> None:
    """Mirrors the real template's actual layout (samplefiles/SampleCodeReview.xlsx):
    a merged title row, a real header row with 'Clause' (not 'Category'), no
    separate Score column (sub-row scores live in 'Avg Points'), category rows
    carrying pre-existing rollup formulas, and the first sub-row under a
    category left with a blank id cell.
    """
    wb = Workbook()
    ws = wb.active
    ws.append(["<Project Name>", None, None, None, None, None, None])
    ws.append(["Clause", None, "Weight", "Avg Points", "Final Points", "% Points", "Remarks"])
    for cell in ws[2]:
        cell.font = Font(bold=True)

    ws.append([1, "Code naming conventions / Code Structure", 1, "=AVERAGE(D3:D4)", "=D3*C3", "=E3/C3", None])
    ws.append([None, "Clear and consistent naming", None, None, None, None, None])
    ws.append([1.2, "Clean structure and formatting", None, None, None, None, None])
    ws.append([None, None, None, None, None, None, None])
    ws.append([None, "General Remarks: placeholder text", None, None, None, None, None])
    ws.append([None, "Reviewers: ", "<reviewer Name>", None, None, None, None])
    ws.append([None, "Dated", None, None, None, None, None])
    wb.save(path)


def test_aggregate_category_scores_computes_mean_and_percent():
    sub_scores = {
        "1.1": {"score": 1, "remark": "Good naming"},
        "1.2": {"score": 0.5, "remark": "Some issues"},
    }
    result = aggregate_category_scores(sub_scores)
    assert result["avg_points"] == 0.75
    assert result["final_points"] == 0.75
    assert result["percent_points"] == 75.0
    assert result["sub_scores"] == sub_scores


def test_aggregate_category_scores_all_none_stays_none():
    sub_scores = {"1.1": {"score": None, "remark": ""}}
    result = aggregate_category_scores(sub_scores)
    assert result["avg_points"] is None
    assert result["final_points"] is None
    assert result["percent_points"] is None


def test_compute_total_score_pct_averages_category_percentages():
    scores_by_category = {
        "1": {"avg_points": 0.9, "final_points": 0.9, "percent_points": 90.0, "sub_scores": {}},
        "2": {"avg_points": 0.6, "final_points": 0.6, "percent_points": 60.0, "sub_scores": {}},
    }
    assert compute_total_score_pct(scores_by_category) == 75.0


def test_compute_total_score_pct_skips_categories_with_no_score():
    scores_by_category = {
        "1": {"avg_points": 1.0, "final_points": 1.0, "percent_points": 100.0, "sub_scores": {}},
        "2": {"avg_points": None, "final_points": None, "percent_points": None, "sub_scores": {}},
    }
    assert compute_total_score_pct(scores_by_category) == 100.0


def test_compute_total_score_pct_returns_none_when_no_category_has_a_score():
    scores_by_category = {
        "1": {"avg_points": None, "final_points": None, "percent_points": None, "sub_scores": {}},
    }
    assert compute_total_score_pct(scores_by_category) is None


def test_populate_scores_writes_sub_row_scores_positionally(tmp_path: Path):
    template_path = tmp_path / "template.xlsx"
    _build_template(template_path)
    wb = load_workbook(template_path)
    ws = wb.active

    category_results = {
        "1": aggregate_category_scores(
            {
                "1.1": {"score": 1, "remark": "Good naming"},
                "1.2": {"score": 0.5, "remark": "Some issues"},
            }
        )
    }
    populate_scores(ws, category_results)

    header_row = [c.value for c in ws[2]]
    assert header_row == ["Clause", None, "Weight", "Avg Points", "Final Points", "% Points", "Remarks"]
    assert ws["A2"].font.bold is True

    # Category row's own rollup formulas are untouched (never written to).
    category_row = ws[3]
    assert category_row[3].value == "=AVERAGE(D3:D4)"
    assert category_row[4].value == "=D3*C3"
    assert category_row[5].value == "=E3/C3"

    # First sub-row (blank id cell) still gets matched positionally as "1.1".
    # Perfect score (1) -- remark is intentionally left blank (self-explanatory).
    sub_row_1_1 = ws[4]
    assert sub_row_1_1[0].value is None  # id cell stays blank, untouched
    assert sub_row_1_1[3].value == 1  # score written into "Avg Points"
    assert sub_row_1_1[6].value is None

    # Second sub-row (labeled 1.2) matched positionally as the category's 2nd
    # sub-criterion. Imperfect score (0.5) -- remark IS written.
    sub_row_1_2 = ws[5]
    assert sub_row_1_2[3].value == 0.5
    assert sub_row_1_2[6].value == "Some issues"


def test_populate_scores_only_writes_remark_when_score_is_not_perfect(tmp_path: Path):
    template_path = tmp_path / "template.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.append(["Clause", None, "Weight", "Avg Points", "Final Points", "% Points", "Remarks"])
    ws.append([1, "Category", 1, "=AVERAGE(D3:D6)", None, None, None])
    ws.append([1.1, "Perfect", None, None, None, None, None])
    ws.append([1.2, "Partial", None, None, None, None, None])
    ws.append([1.3, "Failing", None, None, None, None, None])
    ws.append([1.4, "Not evaluated", None, None, None, None, None])
    wb.save(template_path)
    ws2 = load_workbook(template_path).active

    category_results = {
        "1": aggregate_category_scores(
            {
                "1.1": {"score": 1, "remark": "Perfect naming"},
                "1.2": {"score": 0.5, "remark": "Some issues"},
                "1.3": {"score": 0, "remark": "No structure at all"},
                "1.4": {"score": None, "remark": "Nothing to evaluate here"},
            }
        )
    }
    populate_scores(ws2, category_results)

    assert ws2["D3"].value == 1 and ws2["G3"].value is None  # score 1 -> no remark
    assert ws2["D4"].value == 0.5 and ws2["G4"].value == "Some issues"
    assert ws2["D5"].value == 0 and ws2["G5"].value == "No structure at all"
    assert ws2["D6"].value is None and ws2["G6"].value == "Nothing to evaluate here"


def test_populate_scores_clears_stale_remark_when_score_becomes_perfect(tmp_path: Path):
    # Simulates re-running a review against a template that already has a
    # remark left over from a prior run -- a newly-perfect score must clear
    # it, not just skip writing a new one, or the sheet would show a stale
    # criticism next to a passing score.
    template_path = tmp_path / "template.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.append(["Clause", None, "Weight", "Avg Points", "Final Points", "% Points", "Remarks"])
    ws.append([1, "Category", 1, "=AVERAGE(D3:D3)", None, None, None])
    ws.append([1.1, "Now fixed", None, "Old stale remark from a prior run", None, None, "Old stale remark from a prior run"])
    wb.save(template_path)
    ws2 = load_workbook(template_path).active

    category_results = {
        "1": aggregate_category_scores({"1.1": {"score": 1, "remark": "Looks good now"}}),
    }
    populate_scores(ws2, category_results)

    assert ws2["D3"].value == 1
    assert ws2["G3"].value is None


def test_populate_scores_raises_on_missing_columns(tmp_path: Path):
    template_path = tmp_path / "bad_template.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.append(["Category", "Description"])
    wb.save(template_path)

    ws2 = load_workbook(template_path).active
    try:
        populate_scores(ws2, {})
        assert False, "expected ValueError"
    except ValueError as exc:
        assert "missing" in str(exc).lower()


def test_populate_scores_raises_when_no_header_row_found(tmp_path: Path):
    template_path = tmp_path / "no_header_template.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.append(["Nothing", "Relevant", "Here"])
    wb.save(template_path)

    ws2 = load_workbook(template_path).active
    try:
        populate_scores(ws2, {})
        assert False, "expected ValueError"
    except ValueError as exc:
        assert "header row" in str(exc).lower()


def test_populate_scores_matches_excel_native_numeric_category_id(tmp_path: Path):
    template_path = tmp_path / "numeric_ids_template.xlsx"
    _build_template(template_path)
    ws = load_workbook(template_path).active

    # Category id cell (A3) is already an Excel-native int (1), not a string,
    # per _build_template -- this test exercises that _normalize_id handles it.
    category_results = {
        "1": aggregate_category_scores(
            {
                "1.1": {"score": 0.5, "remark": "Good naming"},
            }
        )
    }
    populate_scores(ws, category_results)

    sub_row_1_1 = ws[4]
    assert sub_row_1_1[3].value == 0.5
    assert sub_row_1_1[6].value == "Good naming"


def test_populate_metadata_fills_title_remarks_reviewer_and_date(tmp_path: Path):
    template_path = tmp_path / "template.xlsx"
    _build_template(template_path)
    ws = load_workbook(template_path).active

    review_date = datetime.date(2026, 7, 24)
    populate_metadata(
        ws,
        project_name="My Cool App",
        general_remarks="Solid overall, exception handling needs work.",
        reviewer_name="Claude",
        review_date=review_date,
    )

    assert ws["A1"].value == "My Cool App"
    assert ws["B7"].value == "General Remarks: Solid overall, exception handling needs work."
    assert ws["C8"].value == "Claude"
    assert ws["C9"].value == review_date


def test_populate_metadata_leaves_unmatched_labels_untouched(tmp_path: Path):
    template_path = tmp_path / "no_metadata_template.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.append(["Clause", None, "Weight", "Avg Points", "Final Points", "% Points", "Remarks"])
    wb.save(template_path)
    ws2 = load_workbook(template_path).active

    # Should not raise even though none of the metadata labels exist in this sheet.
    populate_metadata(ws2, project_name="X", general_remarks="Y", reviewer_name="Claude", review_date=datetime.date.today())
    assert ws2["A1"].value == "Clause"


def test_generate_review_excel_writes_scores_and_metadata_in_one_save(tmp_path: Path):
    template_path = tmp_path / "template.xlsx"
    output_path = tmp_path / "output.xlsx"
    _build_template(template_path)

    category_results = {
        "1": aggregate_category_scores({"1.1": {"score": 0.5, "remark": "Good naming"}}),
    }
    generate_review_excel(
        template_path,
        output_path,
        category_results,
        project_name="My Cool App",
        general_remarks="Looks fine.",
        reviewer_name="Claude",
        review_date=datetime.date(2026, 7, 24),
    )

    wb = load_workbook(output_path)
    ws = wb.active
    assert ws["A1"].value == "My Cool App"
    assert ws["C8"].value == "Claude"
    # openpyxl always reads date-typed cells back from disk as datetime, even
    # though a plain date was written -- expected, not a bug.
    assert ws["C9"].value == datetime.datetime(2026, 7, 24)
    assert ws[4][3].value == 0.5
    assert ws[4][6].value == "Good naming"


def test_generate_review_excel_defaults_reviewer_to_claude_and_date_to_today(tmp_path: Path):
    template_path = tmp_path / "template.xlsx"
    output_path = tmp_path / "output.xlsx"
    _build_template(template_path)

    generate_review_excel(
        template_path,
        output_path,
        {},
        project_name="My Cool App",
        general_remarks="Looks fine.",
    )

    wb = load_workbook(output_path)
    ws = wb.active
    assert ws["C8"].value == "Claude"
    assert ws["C9"].value.date() == datetime.date.today()


def test_populate_scores_against_the_real_sample_template(tmp_path: Path):
    """End-to-end against the actual production template (not a synthetic mimic),
    covering all 5 categories including the ones with typo'd/duplicate/missing
    id labels in column A (e.g. category 4's rows are labeled 4.2, 4.3, 4.3
    instead of 4.1, 4.2, 4.3) -- positional matching must handle this correctly
    since the category's own AVERAGE(...) formula range is 3 rows regardless
    of what the id column says. Also exercises the metadata fields against the
    real title/general-remarks/reviewer/date cells.
    """
    template_path = FIXTURES_DIR / "SampleCodeReview.xlsx"
    output_path = tmp_path / "output.xlsx"

    category_results = {
        "1": aggregate_category_scores({sub_id: {"score": 1, "remark": f"r{sub_id}"} for sub_id in ["1.1", "1.2", "1.3", "1.4", "1.5", "1.6"]}),
        "2": aggregate_category_scores({sub_id: {"score": 0.5, "remark": f"r{sub_id}"} for sub_id in ["2.1", "2.2", "2.3", "2.4"]}),
        "3": aggregate_category_scores({sub_id: {"score": 1, "remark": f"r{sub_id}"} for sub_id in ["3.1", "3.2", "3.3", "3.4"]}),
        "4": aggregate_category_scores({sub_id: {"score": 0, "remark": f"r{sub_id}"} for sub_id in ["4.1", "4.2", "4.3"]}),
        "6": aggregate_category_scores({sub_id: {"score": 1, "remark": f"r{sub_id}"} for sub_id in ["6.1", "6.2", "6.3"]}),
    }
    generate_review_excel(
        template_path,
        output_path,
        category_results,
        project_name="Real App",
        general_remarks="Overall solid, needs test coverage work.",
        reviewer_name="Claude",
        review_date=datetime.date(2026, 7, 24),
    )

    wb = load_workbook(output_path)
    ws = wb.active

    # Category 1 rollup formulas untouched; its 6 sub-rows (4-9) populated in order.
    # Score is a perfect 1 for this category -- remarks intentionally blank.
    assert ws["D3"].value == "=AVERAGE(D4:D9)"
    for row in range(4, 10):
        assert ws.cell(row=row, column=4).value == 1
        assert ws.cell(row=row, column=7).value is None

    # Category 2 (score 0.5, imperfect) -- remarks ARE written.
    assert ws["D11"].value == "=AVERAGE(D12:D15)"
    for row, expected_remark in zip(range(12, 16), ["r2.1", "r2.2", "r2.3", "r2.4"]):
        assert ws.cell(row=row, column=4).value == 0.5
        assert ws.cell(row=row, column=7).value == expected_remark

    # Category 4 (rows 24-26) is labeled 4.2/4.3/4.3 in the real file -- positional
    # matching must still place 4.1/4.2/4.3's scores into rows 24/25/26 in order.
    # Score is 0 (imperfect) -- remarks ARE written.
    assert ws["D23"].value == "=AVERAGE(D24:D26)"
    for row, expected_remark in zip(range(24, 27), ["r4.1", "r4.2", "r4.3"]):
        assert ws.cell(row=row, column=4).value == 0
        assert ws.cell(row=row, column=7).value == expected_remark

    # Category 6 (rows 29-31), last category before the "Total" row. Perfect
    # score again -- remarks intentionally blank.
    assert ws["D28"].value == "=AVERAGE(D29:D31)"
    for row in range(29, 32):
        assert ws.cell(row=row, column=4).value == 1
        assert ws.cell(row=row, column=7).value is None

    # "Total" row and other trailing rows are untouched, not mistaken for a category.
    assert ws["A32"].value == "Total"
    assert ws["C32"].value == "=SUM(C3:C30)"

    # Real metadata cells, per samplefiles/SampleCodeReview.xlsx's actual layout.
    assert ws["A1"].value == "Real App"
    assert ws["B36"].value == "General Remarks: Overall solid, needs test coverage work."
    assert ws["C37"].value == "Claude"
    assert ws["C40"].value == datetime.datetime(2026, 7, 24)


def test_extract_sub_criteria_descriptions_reads_positionally(tmp_path: Path):
    template_path = tmp_path / "template.xlsx"
    _build_template(template_path)
    ws = load_workbook(template_path).active

    categories = {"1": {"name": "Code Structure", "sub_criteria": ["1.1", "1.2"]}}
    descriptions = extract_sub_criteria_descriptions(ws, categories)

    # First sub-row's id cell is blank in _build_template, but its description
    # is still read correctly since matching is positional, not id-based.
    assert descriptions == {
        "1.1": "Clear and consistent naming",
        "1.2": "Clean structure and formatting",
    }


def test_extract_sub_criteria_descriptions_against_the_real_sample_template():
    """Grounds the fix for the reported bug: the LLM prompt only sent bare ids
    like "2.4", never their actual meaning, causing remarks unrelated to the
    real criterion (e.g. a keystore-storage criterion getting an EventBus
    remark). This locks in that the exact real wording is extracted for
    every category, including the mislabeled category 4 rows.
    """
    ws = load_workbook(FIXTURES_DIR / "SampleCodeReview.xlsx").active
    categories = {
        "1": {"name": "x", "sub_criteria": ["1.1", "1.2", "1.3", "1.4", "1.5", "1.6"]},
        "2": {"name": "x", "sub_criteria": ["2.1", "2.2", "2.3", "2.4"]},
        "4": {"name": "x", "sub_criteria": ["4.1", "4.2", "4.3"]},
    }
    descriptions = extract_sub_criteria_descriptions(ws, categories)

    assert descriptions["1.1"] == "Clear and consistent naming conventions"
    assert descriptions["2.4"] == "Keystore information should be stored in env. Or gradle"
    # Category 4's rows are labeled 4.2/4.3/4.3 in the real file, but
    # positional matching still assigns the right description to 4.1/4.2/4.3.
    assert descriptions["4.1"] == "AI usage declared in PR comments along with tool name  (e.g., Copilot, ChatGPT, Azure OpenAI)"
    assert descriptions["4.3"] == "No unexplained or uncommented complex logic, No blind copy-paste"
