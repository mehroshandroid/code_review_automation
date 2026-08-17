import io
import time
import zipfile
from pathlib import Path

from fastapi.testclient import TestClient
from openpyxl import Workbook, load_workbook

import app.api.reviews as reviews_module
from main import app

JACOCO_XML = """<?xml version="1.0" encoding="UTF-8"?>
<report name="test">
    <counter type="INSTRUCTION" missed="10" covered="90"/>
</report>
"""


def _build_zip_bytes() -> bytes:
    buffer = io.BytesIO()
    with zipfile.ZipFile(buffer, "w") as zf:
        zf.writestr(
            "app/build.gradle",
            "apply plugin: 'jacoco'\n"
            "android { compileSdkVersion 34\n defaultConfig { targetSdkVersion 34 } }\n",
        )
        zf.writestr("AndroidManifest.xml", "<manifest />")
        zf.writestr("src/main/java/com/example/MainActivity.java", "class MainActivity {}")
        zf.writestr("build/reports/jacoco/jacocoTestReport.xml", JACOCO_XML)
    return buffer.getvalue()


def _build_ios_zip_bytes() -> bytes:
    buffer = io.BytesIO()
    with zipfile.ZipFile(buffer, "w") as zf:
        zf.writestr("MyApp.xcodeproj/project.pbxproj", "buildSettings = { IPHONEOS_DEPLOYMENT_TARGET = 17.0; SWIFT_VERSION = 5.9; };")
        zf.writestr("Info.plist", "<plist></plist>")
        zf.writestr("MyApp/AppDelegate.swift", "class AppDelegate {}")
    return buffer.getvalue()


def _build_dotnet_zip_bytes() -> bytes:
    buffer = io.BytesIO()
    with zipfile.ZipFile(buffer, "w") as zf:
        zf.writestr("MyApp.sln", "stub")
        zf.writestr(
            "MyApp/MyApp.csproj",
            "<Project><PropertyGroup><TargetFramework>net8.0</TargetFramework></PropertyGroup></Project>",
        )
        zf.writestr("MyApp/Program.cs", "class Program {}")
    return buffer.getvalue()


def _build_xlsx_bytes() -> bytes:
    """Mirrors the real production template's layout (samplefiles/SampleCodeReview.xlsx):
    a title row, a 'Clause' header row, category rows carrying pre-existing rollup
    formulas, and the first sub-row under each category left with a blank id cell.
    Includes categories 1 (6 sub-criteria) and 2 (4 sub-criteria) in full, matching
    CATEGORIES' real counts, since populate_scores consumes rows positionally.
    """
    buffer = io.BytesIO()
    wb = Workbook()
    ws = wb.active
    ws.append(["<Project Name>", None, None, None, None, None, None])
    ws.append(["Clause", None, "Weight", "Avg Points", "Final Points", "% Points", "Remarks"])
    ws.append([1, "Code naming conventions / Code Structure", 1, "=AVERAGE(D4:D9)", "=D3*C3", "=E3/C3", None])
    ws.append([None, "Clear and consistent naming conventions", None, None, None, None, None])
    ws.append([1.2, "Clean structure, formatting, and file organization", None, None, None, None, None])
    ws.append([1.3, "No unused, dead, or commented code", None, None, None, None, None])
    ws.append([1.4, "No compile-time warnings", None, None, None, None, None])
    ws.append([1.5, "No unused dependencies", None, None, None, None, None])
    ws.append([1.6, "Latest compile, target sdk and gradle versions", None, None, None, None, None])
    ws.append([2, "Reliability, Security & Observability", 1, "=AVERAGE(D11:D14)", "=D10*C10", "=E10/C10", None])
    ws.append([2.1, "Proper exception handling", None, None, None, None, None])
    ws.append([2.2, "Centralized logging with correct levels", None, None, None, None, None])
    ws.append([2.3, "No sensitive data stored or logged", None, None, None, None, None])
    ws.append([2.4, "Keystore information stored in env or gradle", None, None, None, None, None])
    wb.save(buffer)
    return buffer.getvalue()


def test_full_review_pipeline_in_stub_mode(monkeypatch, tmp_path: Path):
    monkeypatch.delenv("AZURE_OPENAI_KEY", raising=False)

    import zipfile as zipfile_module
    with zipfile_module.ZipFile(io.BytesIO(_build_zip_bytes())) as zf:
        zf.extractall(tmp_path)
    from app.analyzer.android_analyzer import gather_code_context
    assert "class MainActivity {}" in gather_code_context(tmp_path)

    real_score_category = reviews_module.score_category
    captured_code_snippets = []

    async def _capturing_score_category(provider, category_name, sub_criteria, descriptions, code_snippets, model=None, platform="Android", checklists=None):
        captured_code_snippets.append(code_snippets)
        return await real_score_category(
            provider, category_name, sub_criteria, descriptions, code_snippets, model=model, platform=platform
        )

    monkeypatch.setattr(reviews_module, "score_category", _capturing_score_category)

    async def _fake_check_compile_warnings(zip_path_arg):
        return {"status": "ok", "warning_count": 0, "issues": []}

    monkeypatch.setattr(reviews_module, "check_compile_warnings", _fake_check_compile_warnings)

    with TestClient(app) as client:
        create_response = client.post(
            "/api/reviews",
            files={
                "androidZip": ("project.zip", _build_zip_bytes(), "application/zip"),
                "excelTemplate": (
                    "template.xlsx",
                    _build_xlsx_bytes(),
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                ),
            },
        )
        assert create_response.status_code == 200
        review_id = create_response.json()["review_id"]

        final_state = None
        for _ in range(50):
            progress_response = client.get(f"/api/reviews/{review_id}/progress")
            body = progress_response.json()
            if body["status"] in ("completed", "error"):
                final_state = body
                break
            time.sleep(0.05)

        assert final_state is not None, "review did not finish in time"
        assert final_state["status"] == "completed"
        assert final_state["download_url"] == f"/api/reviews/{review_id}/download"
        assert "total_time_ms" in final_state["stats"]
        assert final_state["test_coverage"] == 90.0
        assert final_state["secrets_found"] == []
        assert final_state["warnings"] == []
        # Stub mode scores every sub-criterion 1 (perfect) across both
        # categories in this fixture, so each category's percent_points is
        # 100.0 and the mean across categories is exactly 100.0.
        assert final_state["total_score_pct"] == 100.0
        assert final_state["compile_status"] == "ok"
        assert final_state["lint_issues"] == []
        assert final_state["project_name"] == "project"

        category_1 = next(c for c in final_state["category_scores"] if c["id"] == "1")
        sub_criteria_by_id = {s["id"]: s for s in category_1["sub_criteria"]}
        # Descriptions come from the xlsx fixture's own text (see _build_xlsx_bytes above).
        assert sub_criteria_by_id["1.1"]["description"] == "Clear and consistent naming conventions"
        assert sub_criteria_by_id["1.4"]["description"] == "No compile-time warnings"
        # Stub mode scores every LLM-scored sub-criterion 1 with the stub placeholder remark.
        assert sub_criteria_by_id["1.1"]["score"] == 1
        assert "placeholder score" in sub_criteria_by_id["1.1"]["remark"]
        # 1.4 comes from the (stubbed) compile-check merge, not the LLM, and keeps its own remark.
        assert sub_criteria_by_id["1.4"]["score"] == 1
        assert sub_criteria_by_id["1.4"]["remark"] == "No Lint warnings or errors found."

        category_2 = next(c for c in final_state["category_scores"] if c["id"] == "2")
        sub_2_1 = next(s for s in category_2["sub_criteria"] if s["id"] == "2.1")
        assert sub_2_1["description"] == "Proper exception handling"
        assert sub_2_1["score"] == 1

        # Proves the runtime wiring end to end: extraction -> gather_code_context ->
        # _run_review -> score_category actually receives the gathered content, not
        # just that gather_code_context works in isolation on a separate directory.
        assert captured_code_snippets, "score_category was never invoked by _run_review"
        assert any("class MainActivity {}" in snippets for snippets in captured_code_snippets)

        download_response = client.get(final_state["download_url"])
        assert download_response.status_code == 200

        workbook = load_workbook(io.BytesIO(download_response.content))
        ws = workbook.active

        # Category row's rollup formula is untouched -- populate_scores never
        # writes to it, since the real template computes it via formula.
        category_1_row = ws[3]
        assert category_1_row[3].value == "=AVERAGE(D4:D9)"

        # First sub-row under category 1 (blank id cell in the fixture, matched
        # positionally as 1.1) gets its stub score written. Stub mode always
        # scores 1 (perfect), so the remark is correctly left blank -- remarks
        # are reserved for imperfect scores.
        sub_1_1_row = ws[4]
        assert sub_1_1_row[3].value == 1
        assert sub_1_1_row[6].value is None

        # Metadata: project name derived from the uploaded zip's filename
        # (this fixture only has the title placeholder, not the General
        # Remarks/Reviewers/Dated cells -- those are covered end-to-end
        # against the real template in test_excel_handler.py).
        assert ws["A1"].value == "project"


async def test_full_review_pipeline_static_mode_scores_1_4_via_stub_llm(monkeypatch):
    monkeypatch.delenv("AZURE_OPENAI_KEY", raising=False)

    async def _fail_if_called(zip_path_arg):
        raise AssertionError("check_compile_warnings must not be called in static mode")

    monkeypatch.setattr(reviews_module, "check_compile_warnings", _fail_if_called)

    with TestClient(app) as client:
        create_response = client.post(
            "/api/reviews",
            files={
                "androidZip": ("project.zip", _build_zip_bytes(), "application/zip"),
                "excelTemplate": (
                    "template.xlsx",
                    _build_xlsx_bytes(),
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                ),
            },
            data={"compileCheckMode": "static"},
        )
        assert create_response.status_code == 200
        review_id = create_response.json()["review_id"]

        final_state = None
        for _ in range(50):
            progress_response = client.get(f"/api/reviews/{review_id}/progress")
            body = progress_response.json()
            if body["status"] in ("completed", "error"):
                final_state = body
                break
            time.sleep(0.05)

        assert final_state is not None, "review did not finish in time"
        assert final_state["status"] == "completed"
        assert final_state["compile_status"] == "skipped"
        assert final_state["lint_issues"] == []

        category_1 = next(c for c in final_state["category_scores"] if c["id"] == "1")
        sub_1_4 = next(s for s in category_1["sub_criteria"] if s["id"] == "1.4")
        assert sub_1_4["description"] == "No compile-time warnings"
        assert sub_1_4["score"] == 1
        assert "placeholder score" in sub_1_4["remark"]


async def test_full_review_pipeline_from_devops_source(monkeypatch):
    monkeypatch.delenv("AZURE_OPENAI_KEY", raising=False)

    async def fake_fetch_repo_zip(repo_url, pat, branch=None):
        return {"status": "ok", "content": _build_zip_bytes(), "message": None}

    async def fake_check_compile_warnings(zip_path_arg):
        return {"status": "ok", "warning_count": 0, "issues": []}

    monkeypatch.setattr(reviews_module, "fetch_repo_zip", fake_fetch_repo_zip)
    monkeypatch.setattr(reviews_module, "check_compile_warnings", fake_check_compile_warnings)

    with TestClient(app) as client:
        create_response = client.post(
            "/api/reviews",
            files={
                "excelTemplate": (
                    "template.xlsx", _build_xlsx_bytes(),
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                ),
            },
            data={
                "devopsRepoUrl": "https://dev.azure.com/myorg/MyProject/_git/my-repo",
                "devopsPat": "fake-pat",
            },
        )
        assert create_response.status_code == 200
        review_id = create_response.json()["review_id"]

        final_state = None
        for _ in range(50):
            progress_response = client.get(f"/api/reviews/{review_id}/progress")
            body = progress_response.json()
            if body["status"] in ("completed", "error"):
                final_state = body
                break
            time.sleep(0.05)

        assert final_state is not None, "review did not finish in time"
        assert final_state["status"] == "completed"
        assert final_state["project_name"] == "my-repo"
        assert final_state["compile_status"] == "ok"

        category_1 = next(c for c in final_state["category_scores"] if c["id"] == "1")
        sub_1_1 = next(s for s in category_1["sub_criteria"] if s["id"] == "1.1")
        assert sub_1_1["score"] == 1


async def test_full_review_pipeline_unsupported_platform_skips_compile_check(monkeypatch):
    # Android, iOS, and .NET all have their own analyzer/compile-check
    # story now; a platform with none of those (e.g. Web (React), not yet
    # supported) must still get compile_status=None ("not applicable"),
    # not attempt either checker, and fall back to android_analyzer for
    # its (unused-for-scoring) structural analysis pass.
    monkeypatch.delenv("AZURE_OPENAI_KEY", raising=False)

    async def _fail_if_called(zip_path_arg):
        raise AssertionError("check_compile_warnings must not be called for an unsupported platform")

    monkeypatch.setattr(reviews_module, "check_compile_warnings", _fail_if_called)

    with TestClient(app) as client:
        create_response = client.post(
            "/api/reviews",
            files={
                "androidZip": ("project.zip", _build_zip_bytes(), "application/zip"),
                "excelTemplate": (
                    "template.xlsx",
                    _build_xlsx_bytes(),
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                ),
            },
            data={"platform": "Web (React)"},
        )
        assert create_response.status_code == 200
        review_id = create_response.json()["review_id"]

        final_state = None
        for _ in range(50):
            progress_response = client.get(f"/api/reviews/{review_id}/progress")
            body = progress_response.json()
            if body["status"] in ("completed", "error"):
                final_state = body
                break
            time.sleep(0.05)

        assert final_state is not None, "review did not finish in time"
        assert final_state["status"] == "completed"
        assert final_state["compile_status"] is None
        assert final_state["lint_issues"] == []

        category_1 = next(c for c in final_state["category_scores"] if c["id"] == "1")
        sub_1_4 = next(s for s in category_1["sub_criteria"] if s["id"] == "1.4")
        # No compile-check exclusion for a non-Android platform -- 1.4 is
        # scored by the (stub-mode) LLM like every other sub-criterion.
        assert sub_1_4["score"] == 1
        assert "placeholder score" in sub_1_4["remark"]


async def test_full_review_pipeline_for_a_real_ios_project(monkeypatch):
    monkeypatch.delenv("AZURE_OPENAI_KEY", raising=False)

    with TestClient(app) as client:
        create_response = client.post(
            "/api/reviews",
            files={
                "androidZip": ("MyApp.zip", _build_ios_zip_bytes(), "application/zip"),
                "excelTemplate": (
                    "template.xlsx", _build_xlsx_bytes(),
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                ),
            },
            data={"platform": "iOS"},
        )
        assert create_response.status_code == 200
        review_id = create_response.json()["review_id"]

        final_state = None
        for _ in range(50):
            progress_response = client.get(f"/api/reviews/{review_id}/progress")
            body = progress_response.json()
            if body["status"] in ("completed", "error"):
                final_state = body
                break
            time.sleep(0.05)

        assert final_state is not None, "review did not finish in time"
        assert final_state["status"] == "completed"
        # No mac_build_agent is reachable in this test environment, so the
        # compile-check gracefully reports "unavailable" -- same fallback
        # Android's own compiler service gets when it can't be reached.
        assert final_state["compile_status"] == "unavailable"

        category_1 = next(c for c in final_state["category_scores"] if c["id"] == "1")
        sub_1_1 = next(s for s in category_1["sub_criteria"] if s["id"] == "1.1")
        assert sub_1_1["score"] == 1


async def test_full_review_pipeline_for_a_real_dotnet_project(monkeypatch):
    monkeypatch.delenv("AZURE_OPENAI_KEY", raising=False)

    with TestClient(app) as client:
        create_response = client.post(
            "/api/reviews",
            files={
                "androidZip": ("MyApp.zip", _build_dotnet_zip_bytes(), "application/zip"),
                "excelTemplate": (
                    "template.xlsx", _build_xlsx_bytes(),
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                ),
            },
            data={"platform": ".NET"},
        )
        assert create_response.status_code == 200
        review_id = create_response.json()["review_id"]

        final_state = None
        for _ in range(50):
            progress_response = client.get(f"/api/reviews/{review_id}/progress")
            body = progress_response.json()
            if body["status"] in ("completed", "error"):
                final_state = body
                break
            time.sleep(0.05)

        assert final_state is not None, "review did not finish in time"
        assert final_state["status"] == "completed"
        # No dotnet-compiler service is reachable in this test environment,
        # so the compile-check gracefully reports "unavailable" -- same
        # fallback every other checker gets when its service can't be
        # reached.
        assert final_state["compile_status"] == "unavailable"

        category_1 = next(c for c in final_state["category_scores"] if c["id"] == "1")
        sub_1_1 = next(s for s in category_1["sub_criteria"] if s["id"] == "1.1")
        assert sub_1_1["score"] == 1


async def test_full_review_pipeline_dotnet_static_mode_skips_build_check(monkeypatch):
    monkeypatch.delenv("AZURE_OPENAI_KEY", raising=False)

    async def _fail_if_called(zip_path_arg):
        raise AssertionError("check_dotnet_build_warnings must not be called in static mode")

    monkeypatch.setattr(reviews_module, "check_dotnet_build_warnings", _fail_if_called)

    with TestClient(app) as client:
        create_response = client.post(
            "/api/reviews",
            files={
                "androidZip": ("MyApp.zip", _build_dotnet_zip_bytes(), "application/zip"),
                "excelTemplate": (
                    "template.xlsx", _build_xlsx_bytes(),
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                ),
            },
            data={"platform": ".NET", "compileCheckMode": "static"},
        )
        assert create_response.status_code == 200
        review_id = create_response.json()["review_id"]

        final_state = None
        for _ in range(50):
            progress_response = client.get(f"/api/reviews/{review_id}/progress")
            body = progress_response.json()
            if body["status"] in ("completed", "error"):
                final_state = body
                break
            time.sleep(0.05)

        assert final_state is not None, "review did not finish in time"
        assert final_state["status"] == "completed"
        assert final_state["compile_status"] == "skipped"
        assert final_state["lint_issues"] == []

        category_1 = next(c for c in final_state["category_scores"] if c["id"] == "1")
        sub_1_4 = next(s for s in category_1["sub_criteria"] if s["id"] == "1.4")
        # Static mode skips the build check -- 1.4 is scored by the
        # (stub-mode) LLM like every other sub-criterion.
        assert sub_1_4["score"] == 1
        assert "placeholder score" in sub_1_4["remark"]
